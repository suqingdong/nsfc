import os
import sys
import json

import openpyxl
from openpyxl.styles import Font, PatternFill
import click
from prettytable import PrettyTable
from simple_loggers import SimpleLogger

from nsfc import DEFAULT_DB, version_info
from nsfc.db.model import Project
from nsfc.db.manager import Manager


__epilog__ = click.style('''\
examples:

\b
# 查看帮助
nsfc query
\b
# 列出可用的查询字段
nsfc query -K
\b
# 输出数量
nsfc query -C
\b
# 按批准年份查询
nsfc query -C -s approval_year 2019
\b
# 按批准年份+学科代码(模糊)
nsfc query -C -s approval_year 2019 -s subject_code "%A%"
\b
# 批准年份也可以是一个区间
nsfc query -C -s approval_year 2015-2019 -s subject_code "%C01%"
\b
# 结果输出为.jl文件
nsfc query -s approval_year 2019 -s subject_code "%C0501%" -o C0501.2019.jl
\b
# 结果输出为xlsx文件
nsfc query -s approval_year 2019 -s subject_code "%C0501%" -o C0501.2019.xlsx -F xlsx
\b
# 限制最大输出条数
nsfc query -L 5 -s approval_year 2019
''', fg='yellow')

@click.command(no_args_is_help=True,
               name='query',
               epilog=__epilog__,
               help='query data from the local database')
@click.option('-d', '--dbfile', help='the database file', default=DEFAULT_DB, show_default=True)

@click.option('-s', '--search', help='the search string, eg. project_id 41950410575', multiple=True, nargs=2)

@click.option('-o', '--outfile', help='the output filename')

@click.option('-F', '--format', help='the format of output',
              type=click.Choice(['json', 'jl', 'tsv', 'xlsx']), default='jl',
              show_choices=True, show_default=True)
@click.option('-K', '--keys', help='list the available keys for query', is_flag=True)
@click.option('-C', '--count', help='just output the out of searching', is_flag=True)
@click.option('-L', '--limit', help='the count of limit of output', type=int)
@click.option('-l', '--log-level', help='the level of logging',
              type=click.Choice(SimpleLogger().level_maps), default='info',
              show_choices=True, show_default=True)
def main(**kwargs):

    logger = SimpleLogger('STATS')
    logger.level = logger.level_maps[kwargs['log_level']]

    logger.info(f'input arguments: {kwargs}')

    dbfile = kwargs['dbfile']
    limit = kwargs['limit']
    outfile = kwargs['outfile']

    if kwargs['keys']:
        table = PrettyTable(['Key', 'Comment', 'Type'])
        for k, v in Project.metadata.tables['project'].columns.items():
            table.add_row([k, v.comment, v.type])
        for field in table._field_names:
            table.align[field] = 'l'
        print(click.style(str(table), fg='cyan'))
        exit(0)

    if not os.path.isfile(dbfile):
        logger.error(f'dbfile not exists! [{dbfile}]')
        baidu = version_info['baidu_data']
        logger.info(f'可通过百度网盘下载需要的数据：{baidu}\n'
                    f'下载完成后可通过-d参数指定数据库文件，也可以拷贝文件到：{DEFAULT_DB}')
        exit(1)

    uri = f'sqlite:///{dbfile}'
    with Manager(uri=uri, echo=False, logger=logger) as m:

        query = m.session.query(Project)

        if kwargs['search']:
            for key, value in kwargs['search']:
                if '%' in value:
                    query = query.filter(Project.__dict__[key].like(value))
                elif key in ('approval_year', ) and not value.isdigit():
                    if '-' in value:
                        min_value, max_value = value.split('-')
                        query = query.filter(Project.__dict__[key] >= min_value)
                        query = query.filter(Project.__dict__[key] <= max_value)
                    else:
                        logger.error('bad approval_year: {value}')
                        exit(1)
                else:
                    query = query.filter(Project.__dict__[key] == value)

        if limit:
            query = query.limit(limit)

        logger.debug(str(query))

        if kwargs['count']:
            logger.info(f'count: {query.count()}')
        elif not query.count():
            logger.warning('no result for your input')
        else:
            if outfile and kwargs['format'] == 'xlsx':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'NSFC-RESULT'
                title = [k for k, v in query.first().__dict__.items() if k != '_sa_instance_state']
                ws.append(title)
                for col, v in enumerate(title,1 ):
                    _ = ws.cell(row=1, column=col, value=v)
                    _.font = Font(color='FFFFFF', bold=True)
                    _.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

                for n, row in enumerate(query):
                    context = [v for k, v in row.__dict__.items() if k != '_sa_instance_state']
                    ws.append(context)

                ws.freeze_panes = 'A2'
                wb.save(outfile)
            else:
                out = open(outfile, 'w') if outfile else sys.stdout
                with out:
                    if kwargs['format'] == 'json':
                        data = [{k: v for k, v in row.__dict__.items() if k != '_sa_instance_state'} for row in query]
                        out.write(json.dumps(data, ensure_ascii=False, indent=2) + '\n')
                    else:
                        for n, row in enumerate(query):
                            context = {k: v for k, v in row.__dict__.items() if k != '_sa_instance_state'}
                            if n == 0 and kwargs['format'] == 'tsv':
                                title = '\t'.join(context.keys())
                                out.write(title + '\n')
                            if kwargs['format'] == 'tsv':
                                line = '\t'.join(map(str, context.values()))
                            else:
                                line = json.dumps(context, ensure_ascii=False)
                            out.write(line + '\n')
            if outfile:
                logger.info(f'save file: {outfile}')


if __name__ == '__main__':
    main()
