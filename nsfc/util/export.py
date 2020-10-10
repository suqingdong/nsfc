# -*- coding=utf-8 -*-
import os
import json
import string

try:
    import cPickle as pickle
except ImportError:
    import pickle

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, colors, Alignment, Border, Side

from . import safe_open


BASE_DIR = os.path.dirname(os.path.realpath(__file__))


class Export(object):
    def __init__(self, data, outfile):
        self.data = data
        self.outfile = outfile

    def to_excel(self, fg=colors.WHITE, bg=colors.BLACK, width=18, size=10):
        """
            保存为Excel格式
        """
        book = openpyxl.Workbook()
        sheet = book.active

        title = self.data[0].keys()
        for column, value in enumerate(title, 1):
            _ = sheet.cell(1, column, value=value)
            sheet.column_dimensions[get_column_letter(column)].width = width
            _.alignment = Alignment(horizontal='left',vertical='center', wrap_text=True)
            _.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            _.font = Font(bold=True, color=fg, size=size)

        for row, line in enumerate(self.data, 2):
            for column, value in enumerate(line.values(), 1):
                _ = sheet.cell(row, column, value=value)
                _.font = Font(size=size)
        book.save(self.outfile)

    def to_tsv(self):
        """
            保存为TSV文本格式(TAB分隔)
        """
        with safe_open(self.outfile, 'w') as out:
            title = self.data[0].keys()
            out.write('\t'.join(title) + '\n')
            for line in self.data:
                out.write('\t'.join(line.values()) + '\n')

    def to_json(self, indent=2):
        """
            保存为JSON文本格式
        """
        with safe_open(self.outfile, 'w') as out:
            json.dump(self.data, out, ensure_ascii=False, indent=indent)

    def to_jsonlines(self):
        """
            保存为JSON LINES文本格式
        """
        with safe_open(self.outfile, 'w') as out:
            for line in self.data:
                out.write(json.dumps(line, ensure_ascii=False) + '\n')

    def to_pickle(self):
        """
            保存为Pickle格式
        """
        with safe_open(self.outfile, 'wb') as out:
            pickle.dump(self.data, out)

    def to_html(self):
        """
            保存为HTML格式，使用在线的DataTables插件
        """
        with safe_open(os.path.join(BASE_DIR, 'index.html')) as f:
            tpl = string.Template(f.read())

        with safe_open(self.outfile, 'w') as out:
            columns = [{'data': each, 'title': each} for each in self.data[0].keys()]
            html = tpl.safe_substitute(COLUMNS=json.dumps(columns), DATA=json.dumps(self.data))
            out.write(html)
